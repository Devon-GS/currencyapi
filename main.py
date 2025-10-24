import currencyapicom
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side
from dotenv import load_dotenv
import os

# jason = {'eur': {'meta': {'last_updated_at': '2025-03-03T23:59:59Z'}, 'data': {'GBP': {'code': 'GBP', 'value': 0.8254537584}, 'MUR': {'code': 'MUR', 'value': 48.6313845297}, 'USD': {'code': 'USD', 'value': 1.0485146702}, 'ZAR': {'code': 'ZAR', 'value': 19.4908428347}}}, 'gbp': {'meta': {'last_updated_at': '2025-03-03T23:59:59Z'}, 'data': {'EUR': {'code': 'EUR', 'value': 1.2114548995}, 'GBP': {'code': 'GBP', 'value': 1}, 'MUR': {'code': 'MUR', 'value': 58.914729059}, 'USD': {'code': 'USD', 'value': 1.2702282344}, 'ZAR': {'code': 'ZAR', 'value': 23.6122770478}}}, 'mur': {'meta': {'last_updated_at': '2025-03-03T23:59:59Z'}, 'data': {'EUR': {'code': 'EUR', 'value': 0.0205628528}, 'GBP': {'code': 'GBP', 'value': 0.0169736841}, 'USD': {'code': 'USD', 'value': 0.0215604528}, 'ZAR': {'code': 'ZAR', 'value': 0.4007873315}}}, 'usd': {'meta': {'last_updated_at': '2025-03-03T23:59:59Z'}, 'data': {'EUR': {'code': 'EUR', 'value': 0.9537300988}, 'GBP': {'code': 'GBP', 'value': 0.7872600946}, 'MUR': {'code': 'MUR', 'value': 46.3812151723}, 'ZAR': {'code': 'ZAR', 'value': 18.5890034624}}}, 'zar': {'meta': {'last_updated_at': '2025-03-03T23:59:59Z'}, 'data': {'EUR': {'code': 'EUR', 'value': 0.0513061446}, 'GBP': {'code': 'GBP', 'value': 0.0423508499}, 'MUR': {'code': 'MUR', 'value': 2.4950888447}, 'USD': {'code': 'USD', 'value': 0.0537952452}}}}

# # Load all environment variables from .env 
try:
    load_dotenv()
except Exception:
    messagebox.showerror(title='.env Loading Error', message=traceback.format_exc())

# Download latest currency exchange rates and save to dic
currencies = {}

try:
    client = currencyapicom.Client(os.environ['CURRENCY_API'])
    currencies['eur'] = client.latest('EUR',currencies=['GBP','MUR','USD','ZAR'])
    currencies['gbp'] = client.latest('GBP',currencies=['EUR','MUR','USD','ZAR'])
    currencies['mur'] = client.latest('MUR',currencies=['EUR','GBP','USD','ZAR'])
    currencies['usd'] = client.latest('USD',currencies=['EUR','GBP','MUR','ZAR'])
    currencies['zar'] = client.latest('ZAR',currencies=['EUR','GBP','MUR','USD'])
except Exception:
    messagebox.showerror(title='Loading Currency Error', message=traceback.format_exc())

# SETUP AND PUTING DATA INTO EXCEL
try:
    # Styling setup
    heading_format = NamedStyle(name="heading_format")
    heading_format.font = Font(bold=True, u='single')
    heading_format.alignment = Alignment(horizontal='center')

    # Open workbook
    cwd = os.getcwd()
    file = filedialog.askopenfilename(title='Assets', initialdir=cwd, filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))

    wb = load_workbook(file)
    ws = wb['Currency Exchange Rates']

    i = 2
    for x in currencies:
        name = x
        date = currencies[x]['meta']['last_updated_at'][:10]
        
        for y in currencies[x]['data']:
            currency_name = y
            value = currencies[name]['data'][y]['value']

            # Create Headings
            ws['A1'] = 'Currency'
            ws['B1'] = 'Last Updated'
            ws['C1'] = 'Value'

            # Input data
            ws[f'A{i}'] = f'{name.upper()}/{currency_name}'
            ws[f'B{i}'] = date
            ws[f'C{i}'] = value

            # Styling
            # ws['A1'].style = heading_format
            # ws['B1'].style = "heading_format"
            # ws['C1'].style = "heading_format"

            ws.column_dimensions['A'].width = 9.30
            ws.column_dimensions['B'].width = 12.08
            ws.column_dimensions['C'].width = 11.97

            i += 1 

    # # Save the file
    wb.save(file)
    wb.close()

#     os.system(f'start "EXCEL.EXE" "{file}"')
except InvalidFileException:
    messagebox.showerror(title='Excel Error', message="No File Selected or Improper File Format Selected")
except Exception as error:
    messagebox.showerror(title='Excel Error', message=error)